import openpyxl
import re

# Load the Excel file
wb = openpyxl.load_workbook('C:/Users/Admin/Downloads/kv22.xlsx')
ws = wb.active

print(f"=== DETAILED ANALYSIS ===\n")

# Get headers
headers = [cell.value for cell in ws[1]]
print(f"Headers: {headers}\n")

# Track issues
missing_cccd_rows = []
duplicate_cccd = {}
cccd_column_idx = None

# First, identify which column contains CCCD
for col_idx, header in enumerate(headers, 1):
    if header and ('cccd' in str(header).lower() or 'cmnd' in str(header).lower() or 'căn cước' in str(header).lower()):
        cccd_column_idx = col_idx
        print(f"CCCD column found at index {col_idx}: {header}\n")
        break

# If not found by header, scan first few rows
if not cccd_column_idx:
    print("CCCD column not found by header, scanning data...\n")
    for col_idx in range(1, ws.max_column + 1):
        sample_values = [ws.cell(row=r, column=col_idx).value for r in range(2, min(10, ws.max_row + 1))]
        if any(v and re.match(r'^\d{9,12}$', str(v).strip()) for v in sample_values):
            cccd_column_idx = col_idx
            print(f"CCCD column detected at index {col_idx}\n")
            break

if not cccd_column_idx:
    print("ERROR: Could not identify CCCD column!\n")
else:
    # Analyze each row
    for row_idx in range(2, ws.max_row + 1):
        row_data = [cell.value for cell in ws[row_idx]]
        
        # Skip completely empty rows
        if all(cell is None or str(cell).strip() == '' for cell in row_data):
            continue
        
        # Get CCCD value
        cccd_value = ws.cell(row=row_idx, column=cccd_column_idx).value
        
        if not cccd_value or not re.match(r'^\d{9,12}$', str(cccd_value).strip()):
            missing_cccd_rows.append({
                'row': row_idx,
                'name': row_data[1] if len(row_data) > 1 else 'N/A',
                'cccd': cccd_value,
                'full_data': row_data[:5]
            })
        else:
            cccd = str(cccd_value).strip()
            if cccd in duplicate_cccd:
                duplicate_cccd[cccd].append(row_idx)
            else:
                duplicate_cccd[cccd] = [row_idx]

    # Print missing CCCD details
    print(f"=== ROWS MISSING VALID CCCD ({len(missing_cccd_rows)}) ===")
    for item in missing_cccd_rows[:10]:  # Show first 10
        print(f"Row {item['row']}: Name={item['name']}, CCCD={item['cccd']}")
    if len(missing_cccd_rows) > 10:
        print(f"... and {len(missing_cccd_rows) - 10} more\n")
    
    # Print duplicate CCCD details
    duplicates = {k: v for k, v in duplicate_cccd.items() if len(v) > 1}
    print(f"\n=== DUPLICATE CCCDs ({len(duplicates)}) ===")
    for cccd, rows in list(duplicates.items())[:10]:  # Show first 10
        names = [ws.cell(row=r, column=2).value for r in rows]
        print(f"CCCD {cccd} in rows {rows}: {names}")
    if len(duplicates) > 10:
        print(f"... and {len(duplicates) - 10} more\n")
    
    # Calculate expected upload
    unique_valid_cccd = len([k for k, v in duplicate_cccd.items() if len(v) == 1])
    duplicate_kept = len(duplicates)  # Only first occurrence kept
    
    print(f"\n=== UPLOAD CALCULATION ===")
    print(f"Total data rows: {ws.max_row - 1}")
    print(f"Rows with missing/invalid CCCD: {len(missing_cccd_rows)}")
    print(f"Unique valid CCCDs: {unique_valid_cccd}")
    print(f"Duplicate CCCDs (only 1st kept): {duplicate_kept}")
    print(f"Expected uploaded: {unique_valid_cccd + duplicate_kept}")
    print(f"User expects: 1614")
    print(f"Actual uploaded: 1584")
