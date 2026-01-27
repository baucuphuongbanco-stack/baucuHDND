import openpyxl
import re
import sys
import json

# Force UTF-8 encoding for stdout
sys.stdout.reconfigure(encoding='utf-8')

# Load the Excel file
wb = openpyxl.load_workbook('C:/Users/Admin/Downloads/kv22.xlsx')
ws = wb.active

failed_rows = []
total_data_rows = 0

ignored_keywords = [
    'DANH SÁCH CỬ TRI',
    'Họ và tên',
    'Ngày sinh',
    'Tổng số',
    'Người lập biểu',
    'UBND',
    'Danh sách này được lập',
    'Cử tri tham gia bầu cử',
    'xã: 1614 người',
]

for row_idx in range(2, ws.max_row + 1):
    row_data = [cell.value for cell in ws[row_idx]]
    line = " ".join([str(c).strip() for c in row_data if c is not None])
    
    if not line.strip(): 
        continue
        
    # Check for keywords to ignore (header/footer)
    is_junk = False
    for kw in ignored_keywords:
        if kw.lower() in line.lower():
            is_junk = True
            break
    if is_junk:
        continue

    total_data_rows += 1

    # 1. CCCD Match
    cccd_match = re.search(r'\b\d{9,12}\b', line)
    
    if not cccd_match:
        failed_rows.append({
            'row': row_idx,
            'reason': 'Không tìm thấy số CCCD hợp lệ (9-12 số)',
            'content': line[:200]
        })
        continue
        
    cccd = cccd_match.group(0)
    cccd_index = line.find(cccd)
    
    # 2. Name Extraction check
    pre_cccd = line[:cccd_index].strip()
    
    if len(pre_cccd) < 2:
         failed_rows.append({
            'row': row_idx,
            'reason': 'Không tìm thấy Họ Tên (trước số CCCD)',
            'content': line[:200]
        })
         continue

# Output JSON result
result = {
    "total_rows": total_data_rows,
    "failed_count": len(failed_rows),
    "failed_rows": failed_rows
}

print(f"Total rows in Excel: {total_data_rows}")

# Audit Logic
# 1. Total Excel Data Rows (excluding junk/headers)
# 2. How many are 'Clean' (Valid CCCD + Name)
# 3. How many are 'Failed' (Force Addable but missing data)
# 4. How many were Skipped silently?

# Re-scan for skipped rows
skipped_rows = []
valid_rows_count = 0

for row_idx in range(2, ws.max_row + 1):
    row_data = [cell.value for cell in ws[row_idx]]
    line = " ".join([str(c).strip() for c in row_data if c is not None])
    if not line.strip(): continue

    # Check Junk
    is_junk = False
    for kw in ignored_keywords:
        if kw.lower() in line.lower():
            is_junk = True
            break
    if is_junk: continue
    
    # Check if "Skipped" (len <= 20)
    if len(line.strip()) <= 20:
        skipped_rows.append({
            'row': row_idx,
            'content': line.strip()
        })
        continue

    # Check Valid vs Failed
    cccd_match = re.search(r'\b\d{9,12}\b', line)
    if cccd_match:
         valid_rows_count += 1
    else:
         # This is the "Force Addable" group
         pass

result['skipped_count'] = len(skipped_rows)
result['skipped_rows'] = skipped_rows
result['valid_count'] = valid_rows_count

with open('analysis_result_final.json', 'w', encoding='utf-8') as f:
    json.dump(result, f, ensure_ascii=False, indent=2)

print(f"Valid Rows: {valid_rows_count}")
print(f"Failed Rows (Force Addable): {len(failed_rows)}")
print(f"Skipped Rows (Short/Empty): {len(skipped_rows)}")
print("Analysis complete. Saved to analysis_result_final.json")
