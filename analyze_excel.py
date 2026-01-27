import openpyxl
import re

# Load the Excel file
wb = openpyxl.load_workbook('C:/Users/Admin/Downloads/kv22.xlsx')
ws = wb.active

print(f"=== FILE ANALYSIS ===")
print(f"Total rows (including header): {ws.max_row}")
print(f"Total data rows: {ws.max_row - 1}")
print(f"Total columns: {ws.max_column}")

# Get headers
headers = [cell.value for cell in ws[1]]
print(f"\nColumn headers: {headers}")

# Analyze data
missing_cccd_count = 0
empty_rows = 0
duplicate_cccd = {}
rows_with_data = 0

for row_idx in range(2, ws.max_row + 1):
    row_data = [cell.value for cell in ws[row_idx]]
    
    # Check if row is completely empty
    if all(cell is None or str(cell).strip() == '' for cell in row_data):
        empty_rows += 1
        continue
    
    rows_with_data += 1
    
    # Find CCCD column (usually contains 9-12 digit numbers)
    cccd_found = False
    for cell_value in row_data:
        if cell_value and re.match(r'^\d{9,12}$', str(cell_value).strip()):
            cccd = str(cell_value).strip()
            cccd_found = True
            
            # Track duplicates
            if cccd in duplicate_cccd:
                duplicate_cccd[cccd].append(row_idx)
            else:
                duplicate_cccd[cccd] = [row_idx]
            break
    
    if not cccd_found:
        missing_cccd_count += 1
        print(f"\nRow {row_idx} missing CCCD: {row_data[:5]}...")

print(f"\n=== SUMMARY ===")
print(f"Total rows with data: {rows_with_data}")
print(f"Empty rows: {empty_rows}")
print(f"Rows missing CCCD: {missing_cccd_count}")
print(f"Duplicate CCCDs: {sum(1 for v in duplicate_cccd.values() if len(v) > 1)}")

if sum(1 for v in duplicate_cccd.values() if len(v) > 1) > 0:
    print(f"\nDuplicate CCCD details:")
    for cccd, rows in duplicate_cccd.items():
        if len(rows) > 1:
            print(f"  CCCD {cccd} appears in rows: {rows}")

print(f"\n=== EXPECTED vs ACTUAL ===")
print(f"Expected records: 1614")
print(f"Actual uploadable records: {rows_with_data - missing_cccd_count}")
print(f"Missing: {1614 - (rows_with_data - missing_cccd_count)}")
